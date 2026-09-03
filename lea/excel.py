"""Robust Excel reading: header detection, fingerprinting, normalisation.

Seller-provided workbooks do not look like tables. Headers sit several rows
down and indented a column or two, footnotes float above them, totals rows sit
below the data, and a dimension is sometimes encoded as the sheet tab name.
Nothing here assumes A1 is a header.
"""

from __future__ import annotations

import hashlib
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

# A row is treated as the header if it has at least this many non-empty cells
# and the rows beneath it are wider than it is sparse.
MIN_HEADER_CELLS = 2
MAX_HEADER_SEARCH_ROWS = 15


@dataclass
class SheetProbe:
    """What we can tell about a sheet before committing to a parse."""

    sheet_name: str
    header_row: int | None
    header: list[str]
    first_data_row: int | None
    n_rows: int
    n_cols: int

    @property
    def normalised_header(self) -> list[str]:
        return [normalise(h) for h in self.header if h]


def content_hash(path: str | Path) -> str:
    """SHA-256 of the file bytes — the idempotency key.

    Hashing content rather than filename means "Copy of census.xlsx" and
    "census (1).xlsx" are correctly recognised as the same load.
    """
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalise(value: Any) -> str:
    """Collapse a header cell to a comparable token.

    'Dept / Title', 'DEPT/TITLE', and 'Dept  /  Title\n' all become
    'dept_title', so a firm renaming a column with different capitalisation or
    stray whitespace does not read as schema drift.
    """
    if value is None:
        return ""
    text = unicodedata.normalize("NFKD", str(value))
    text = text.replace(" ", " ").strip().lower()
    text = re.sub(r"\(.*?\)", " ", text)  # drop parenthetical hints
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def probe_workbook(path: str | Path) -> list[SheetProbe]:
    """Inspect every sheet without parsing it into a target schema."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    probes = []
    try:
        for sheet in workbook.worksheets:
            probes.append(_probe_sheet(sheet))
    finally:
        workbook.close()
    return probes


def _probe_sheet(sheet) -> SheetProbe:
    rows = []
    for i, row in enumerate(sheet.iter_rows(max_row=MAX_HEADER_SEARCH_ROWS, values_only=True)):
        rows.append(row)

    header_row, header = _detect_header(rows)
    n_cols = max((len(r) for r in rows), default=0)

    return SheetProbe(
        sheet_name=sheet.title,
        header_row=header_row,
        header=header,
        first_data_row=(header_row + 1) if header_row is not None else None,
        n_rows=sheet.max_row or 0,
        n_cols=n_cols,
    )


def _detect_header(rows: list[tuple]) -> tuple[int | None, list[str]]:
    """Pick the header row: the widest early row that looks like labels.

    Scoring rewards rows with many non-empty, mostly-textual cells. A totals
    row or a data row scores lower because it is numeric-heavy; a title banner
    scores lower because it has one populated cell.
    """
    best_index, best_score, best_row = None, 0.0, []

    for index, row in enumerate(rows):
        cells = list(row)
        filled = [c for c in cells if c not in (None, "")]
        if len(filled) < MIN_HEADER_CELLS:
            continue

        textual = sum(1 for c in filled if isinstance(c, str))
        score = len(filled) * (textual / len(filled)) ** 2

        if score > best_score:
            best_index, best_score, best_row = index, score, cells

    if best_index is None:
        return None, []

    # Trim trailing empties but keep interior blanks so column positions hold.
    header = [("" if c is None else str(c).strip()) for c in best_row]
    while header and not header[-1]:
        header.pop()
    return best_index, header


def fingerprint(probe: SheetProbe) -> str:
    """A stable id for a sheet's shape, used to look up a saved recipe.

    Built from the sorted set of normalised header tokens, so reordering
    columns does not change the fingerprint but renaming or adding one does.
    """
    tokens = sorted(set(t for t in probe.normalised_header if t))
    return hashlib.sha256("|".join(tokens).encode()).hexdigest()[:16]


def read_rows(
    path: str | Path,
    sheet_name: str,
    header_row: int,
    max_rows: int | None = None,
) -> tuple[list[str], list[list[Any]]]:
    """Read a sheet as (header, rows), starting after the header row."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        header: list[str] = []
        data: list[list[Any]] = []

        for index, row in enumerate(sheet.iter_rows(values_only=True)):
            if index < header_row:
                continue
            if index == header_row:
                header = [("" if c is None else str(c).strip()) for c in row]
                continue
            if all(c in (None, "") for c in row):
                continue
            data.append(list(row))
            if max_rows and len(data) >= max_rows:
                break
    finally:
        workbook.close()

    return header, data


# --- value coercion -------------------------------------------------------
# Excel hands back whatever the author typed. These coercions are deliberately
# strict about returning None rather than guessing: a silently wrong number is
# far more damaging in a valuation model than a visible gap.

_PLACEHOLDER_DATES = {(1, 1)}  # Jan 1 is the classic "we don't know" sentinel
_EXCEL_ERRORS = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!"}


def to_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text in _EXCEL_ERRORS:
        return None
    return text


def to_number(value: Any) -> float | None:
    """Parse a number, tolerating currency symbols, commas, and parens."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text or text in _EXCEL_ERRORS:
        return None

    negative = text.startswith("(") and text.endswith(")")
    text = re.sub(r"[^0-9.\-]", "", text.strip("()"))
    if not text or text in {"-", ".", "-."}:
        return None

    try:
        number = float(text)
    except ValueError:
        return None
    return -number if negative else number


def to_int(value: Any) -> int | None:
    number = to_number(value)
    return None if number is None else int(round(number))


def to_date(value: Any) -> tuple[date | None, bool]:
    """Parse a date. Returns (date, is_placeholder).

    Dates are the messiest field in seller data: real datetimes, US-format
    strings, and Jan-1 sentinels standing in for "sometime that year" all
    arrive in the same column. The placeholder flag lets downstream tenure
    calculations exclude fake precision instead of treating 1999-01-01 as a
    real hire date.
    """
    if value is None:
        return None, False

    parsed: date | None = None
    if isinstance(value, datetime):
        parsed = value.date()
    elif isinstance(value, date):
        parsed = value
    else:
        text = str(value).strip()
        if not text or text in _EXCEL_ERRORS:
            return None, False
        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%y", "%Y/%m/%d"):
            try:
                parsed = datetime.strptime(text, fmt).date()
                break
            except ValueError:
                continue

    if parsed is None:
        return None, False
    return parsed, (parsed.month, parsed.day) in _PLACEHOLDER_DATES


def looks_like_total_row(row: list[Any], label_columns: int = 3) -> bool:
    """Detect a trailing totals row so it is not ingested as a record."""
    labels = " ".join(str(c).lower() for c in row[:label_columns] if c is not None)
    return any(word in labels for word in ("total", "sum", "grand", "subtotal"))
